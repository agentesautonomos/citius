import os
import io
import json
import re
import secrets
import httpx
import redis
from fastapi import FastAPI, Request, Depends, HTTPException, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from anthropic import Anthropic
from datetime import datetime, timedelta

# ============================================================
# CONFIGURACAO
# ============================================================
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY")
ZAPI_INSTANCE_ID  = os.environ.get("ZAPI_INSTANCE_ID")
ZAPI_TOKEN        = os.environ.get("ZAPI_TOKEN")
ZAPI_CLIENT_TOKEN = os.environ.get("ZAPI_CLIENT_TOKEN")
REDIS_URL         = os.environ.get("REDIS_URL")
ADMIN_USER        = os.environ.get("ADMIN_USER", "admin")
ADMIN_PASS        = os.environ.get("ADMIN_PASS", "admin123")
BASE_URL          = os.environ.get("BASE_URL", "https://SEU-DOMINIO.up.railway.app")
AGENT_NAME        = os.environ.get("AGENT_NAME", "PrimeiraMente")
AGENT_MODEL       = os.environ.get("AGENT_MODEL", "claude-haiku-4-5-20251001")
GROQ_API_KEY      = os.environ.get("GROQ_API_KEY")
MP_ACCESS_TOKEN   = os.environ.get("MP_ACCESS_TOKEN")
MP_PUBLIC_KEY     = os.environ.get("MP_PUBLIC_KEY")
MP_PLAN_ID        = os.environ.get("MP_PLAN_ID", "bc34d81de9ba466b8d2693d1a134871c")
LINK_PAGAMENTO    = os.environ.get("LINK_PAGAMENTO", "")  # link externo opcional

AGENT_PROMPT_PADRAO = """Voce e um assistente prestativo e simpatico.
Responda de forma clara, direta e em portugues.
No WhatsApp, seja breve — uma ideia por mensagem, no maximo."""

client   = Anthropic(api_key=ANTHROPIC_API_KEY)
app      = FastAPI()
security = HTTPBasic()

# ============================================================
# REDIS
# ============================================================
r = redis.from_url(REDIS_URL, decode_responses=True)

# ============================================================
# AUTENTICACAO ADMIN
# ============================================================
def verificar_admin(credentials: HTTPBasicCredentials = Depends(security)):
    usuario_ok = secrets.compare_digest(credentials.username.encode(), ADMIN_USER.encode())
    senha_ok   = secrets.compare_digest(credentials.password.encode(), ADMIN_PASS.encode())
    if not (usuario_ok and senha_ok):
        raise HTTPException(
            status_code=401,
            detail="Acesso negado",
            headers={"WWW-Authenticate": "Basic"}
        )
    return credentials.username

# ============================================================
# PROMPT — salvo no Redis, editavel pelo painel
# ============================================================
PROMPT_KEY = "config:agent_prompt"

def obter_prompt() -> str:
    prompt = r.get(PROMPT_KEY)
    return prompt if prompt else AGENT_PROMPT_PADRAO

def salvar_prompt(prompt: str):
    r.set(PROMPT_KEY, prompt)

# ============================================================
# ARQUIVOS DE REFERENCIA
# ============================================================
ARQUIVO_PREFIX = "config:arquivo:"

def listar_arquivos() -> list:
    chaves = r.keys(f"{ARQUIVO_PREFIX}*")
    arquivos = []
    for chave in sorted(chaves):
        nome = chave.replace(ARQUIVO_PREFIX, "")
        tamanho = len(r.get(chave) or "")
        arquivos.append({"nome": nome, "tamanho": tamanho})
    return arquivos

def obter_arquivo(nome: str) -> str | None:
    return r.get(f"{ARQUIVO_PREFIX}{nome}")

def salvar_arquivo(nome: str, conteudo: str):
    r.set(f"{ARQUIVO_PREFIX}{nome}", conteudo[:20000])

def apagar_arquivo(nome: str):
    r.delete(f"{ARQUIVO_PREFIX}{nome}")

def injetar_arquivos_no_prompt(prompt: str) -> str:
    referencias = re.findall(r'\[([a-zA-Z0-9_\-]+)\]', prompt)
    for nome in referencias:
        conteudo = obter_arquivo(nome)
        if conteudo:
            prompt = prompt.replace(f"[{nome}]", f"\n\n=== CONTEUDO DE '{nome}' ===\n{conteudo}\n=== FIM DE '{nome}' ===\n")
    return prompt

# ============================================================
# ASSINATURAS
# ============================================================
ASSINATURA_PREFIX  = "assinatura:"
CONSULTA_PREFIX    = "consulta:"
AGENDAMENTO_PREFIX = "agendamento:"
BLOQUEIO_PREFIX    = "bloqueio:"
CONFIG_AGENDA_KEY  = "config:agenda"

VALOR_CONSULTA     = 350.00
DURACAO_CONSULTA   = 45  # minutos

HORARIOS_PADRAO = ["08:00","09:00","10:00","11:00","14:00","15:00","16:00","17:00","18:00"]
DIAS_SEMANA_UTEIS = [0,1,2,3,4]  # segunda a sexta

# ============================================================
# AGENDA — configuracao
# ============================================================

def obter_config_agenda() -> dict:
    dados = r.get(CONFIG_AGENDA_KEY)
    if not dados:
        return {"horarios": HORARIOS_PADRAO, "dias_uteis": DIAS_SEMANA_UTEIS, "valor": VALOR_CONSULTA}
    return json.loads(dados)

def salvar_config_agenda(config: dict):
    r.set(CONFIG_AGENDA_KEY, json.dumps(config))

# ============================================================
# AGENDA — agendamentos
# ============================================================

def criar_agendamento_id(data: str, hora: str) -> str:
    """Gera chave unica: data_hora ex: 2026-03-10_14:00"""
    return f"{data}_{hora}"

def obter_agendamento(ag_id: str) -> dict | None:
    dados = r.get(f"{AGENDAMENTO_PREFIX}{ag_id}")
    return json.loads(dados) if dados else None

def salvar_agendamento(ag: dict):
    ag_id = criar_agendamento_id(ag["data"], ag["hora"])
    r.set(f"{AGENDAMENTO_PREFIX}{ag_id}", json.dumps(ag))

def cancelar_agendamento(ag_id: str):
    dados = r.get(f"{AGENDAMENTO_PREFIX}{ag_id}")
    if dados:
        ag = json.loads(dados)
        ag["status"] = "cancelado"
        r.set(f"{AGENDAMENTO_PREFIX}{ag_id}", json.dumps(ag))

def apagar_agendamento(ag_id: str):
    r.delete(f"{AGENDAMENTO_PREFIX}{ag_id}")

def listar_agendamentos(apenas_futuros: bool = False) -> list:
    chaves = r.keys(f"{AGENDAMENTO_PREFIX}*")
    resultado = []
    hoje = datetime.now().date()
    for chave in sorted(chaves):
        dados = json.loads(r.get(chave) or "{}")
        if apenas_futuros:
            try:
                data_ag = datetime.strptime(dados.get("data",""), "%Y-%m-%d").date()
                if data_ag < hoje:
                    continue
            except Exception:
                pass
        resultado.append(dados)
    return sorted(resultado, key=lambda x: (x.get("data",""), x.get("hora","")))

def horario_disponivel(data: str, hora: str) -> bool:
    """Verifica se horario esta livre (sem agendamento confirmado e sem bloqueio)."""
    ag_id = criar_agendamento_id(data, hora)
    ag = obter_agendamento(ag_id)
    if ag and ag.get("status") in ("confirmado", "pendente"):
        return False
    if r.get(f"{BLOQUEIO_PREFIX}{data}_{hora}"):
        return False
    if r.get(f"{BLOQUEIO_PREFIX}{data}_dia"):
        return False
    return True

def bloquear_horario(data: str, hora: str):
    r.set(f"{BLOQUEIO_PREFIX}{data}_{hora}", "1")

def desbloquear_horario(data: str, hora: str):
    r.delete(f"{BLOQUEIO_PREFIX}{data}_{hora}")

def bloquear_dia(data: str):
    r.set(f"{BLOQUEIO_PREFIX}{data}_dia", "1")

def desbloquear_dia(data: str):
    r.delete(f"{BLOQUEIO_PREFIX}{data}_dia")

def eh_dia_bloqueado(data: str) -> bool:
    return bool(r.get(f"{BLOQUEIO_PREFIX}{data}_dia"))

def obter_slots_disponiveis(data: str) -> list:
    """Retorna lista de horarios disponiveis para uma data."""
    config = obter_config_agenda()
    horarios = config.get("horarios", HORARIOS_PADRAO)
    try:
        dt = datetime.strptime(data, "%Y-%m-%d")
        if dt.weekday() not in config.get("dias_uteis", DIAS_SEMANA_UTEIS):
            return []
        if dt.date() < datetime.now().date():
            return []
    except Exception:
        return []
    if eh_dia_bloqueado(data):
        return []
    return [h for h in horarios if horario_disponivel(data, h)]

def gerar_proximas_datas(dias: int = 30) -> list:
    """Gera lista de datas com slots disponíveis nos proximos N dias."""
    config = obter_config_agenda()
    dias_uteis = config.get("dias_uteis", DIAS_SEMANA_UTEIS)
    resultado = []
    hoje = datetime.now().date()
    for i in range(1, dias + 1):
        data = hoje + timedelta(days=i)
        if data.weekday() not in dias_uteis:
            continue
        data_str = data.strftime("%Y-%m-%d")
        slots = obter_slots_disponiveis(data_str)
        if slots:
            resultado.append({"data": data_str, "slots": slots,
                              "dia_semana": ["Seg","Ter","Qua","Qui","Sex","Sab","Dom"][data.weekday()],
                              "data_br": data.strftime("%d/%m/%Y")})
    return resultado

def obter_assinatura(telefone: str) -> dict:
    dados = r.get(f"{ASSINATURA_PREFIX}{telefone}")
    if not dados:
        return {"status": "freemium", "plano": "freemium", "telefone": telefone}
    return json.loads(dados)

def salvar_assinatura(telefone: str, dados: dict):
    r.set(f"{ASSINATURA_PREFIX}{telefone}", json.dumps(dados))

def eh_premium(telefone: str) -> bool:
    assinatura = obter_assinatura(telefone)
    if assinatura.get("status") != "ativo":
        return False
    # Verifica vencimento se existir
    expira = assinatura.get("expira")
    if expira:
        try:
            if datetime.fromisoformat(expira) < datetime.now():
                return False
        except Exception:
            pass
    return True

def listar_assinaturas() -> list:
    chaves = r.keys(f"{ASSINATURA_PREFIX}*")
    resultado = []
    for chave in sorted(chaves):
        dados = json.loads(r.get(chave) or "{}")
        resultado.append(dados)
    return resultado

def registrar_interesse_consulta(telefone: str, nome: str):
    dados = {
        "telefone": telefone,
        "nome": nome,
        "data": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "atendido": False
    }
    r.set(f"{CONSULTA_PREFIX}{telefone}", json.dumps(dados))

def listar_consultas() -> list:
    chaves = r.keys(f"{CONSULTA_PREFIX}*")
    resultado = []
    for chave in sorted(chaves):
        dados = json.loads(r.get(chave) or "{}")
        resultado.append(dados)
    return resultado

def marcar_consulta_atendida(telefone: str):
    dados = json.loads(r.get(f"{CONSULTA_PREFIX}{telefone}") or "{}")
    dados["atendido"] = True
    r.set(f"{CONSULTA_PREFIX}{telefone}", json.dumps(dados))

# ============================================================
# HISTORICO COM REDIS
# ============================================================
HISTORICO_LIMITE = 40

def obter_historico(telefone: str) -> list:
    dados = r.get(f"historico:{telefone}")
    if not dados:
        return []
    return json.loads(dados)[-HISTORICO_LIMITE:]

def salvar_historico(telefone: str, historico: list):
    r.set(f"historico:{telefone}", json.dumps(historico))

def salvar_mensagem(telefone: str, role: str, conteudo: str):
    historico = obter_historico(telefone)
    historico.append({"role": role, "content": conteudo})
    salvar_historico(telefone, historico)

# ============================================================
# PROCESSAMENTO DE MIDIA
# ============================================================

async def transcrever_audio(url_audio: str) -> str:
    if not GROQ_API_KEY:
        return "[Audio recebido, mas GROQ_API_KEY nao configurada]"
    try:
        async with httpx.AsyncClient(timeout=60) as http:
            r_audio = await http.get(url_audio)
            conteudo = r_audio.content
            response = await http.post(
                "https://api.groq.com/openai/v1/audio/transcriptions",
                headers={"Authorization": f"Bearer {GROQ_API_KEY}"},
                files={"file": ("audio.ogg", conteudo, "audio/ogg")},
                data={"model": "whisper-large-v3", "language": "pt"}
            )
            if response.status_code == 200:
                texto = response.json().get("text", "")
                return f"[Audio transcrito]: {texto}"
            return "[Nao foi possivel transcrever o audio]"
    except Exception as e:
        print(f"ERRO ao transcrever audio: {e}")
        return "[Erro ao processar audio]"


async def extrair_texto_pdf(url_arquivo: str) -> str:
    try:
        import pdfplumber
        async with httpx.AsyncClient(timeout=60) as http:
            r_arquivo = await http.get(url_arquivo)
            conteudo = r_arquivo.content
        with pdfplumber.open(io.BytesIO(conteudo)) as pdf:
            paginas = []
            for i, pagina in enumerate(pdf.pages[:20]):
                texto = pagina.extract_text()
                if texto:
                    paginas.append(f"[Pagina {i+1}]\n{texto}")
        texto_completo = "\n\n".join(paginas)
        return f"[Conteudo do PDF enviado pelo usuario]:\n{texto_completo[:8000]}"
    except Exception as e:
        print(f"ERRO ao ler PDF: {e}")
        return "[Nao foi possivel ler o PDF]"


async def extrair_texto_excel(url_arquivo: str) -> str:
    try:
        import openpyxl
        async with httpx.AsyncClient(timeout=60) as http:
            r_arquivo = await http.get(url_arquivo)
            conteudo = r_arquivo.content
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
        linhas_total = []
        for nome_aba in wb.sheetnames[:3]:
            ws = wb[nome_aba]
            linhas_total.append(f"[Aba: {nome_aba}]")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 100:
                    linhas_total.append("... (mais linhas omitidas)")
                    break
                linha = " | ".join(str(c) if c is not None else "" for c in row)
                if linha.strip():
                    linhas_total.append(linha)
        texto_completo = "\n".join(linhas_total)
        return f"[Conteudo da planilha enviada pelo usuario]:\n{texto_completo[:8000]}"
    except Exception as e:
        print(f"ERRO ao ler Excel: {e}")
        return "[Nao foi possivel ler a planilha]"


async def processar_midia(dados: dict) -> str | None:
    audio = dados.get("audio", {})
    if audio and audio.get("audioUrl"):
        return await transcrever_audio(audio["audioUrl"])
    documento = dados.get("document", {})
    if documento:
        url      = documento.get("documentUrl", "")
        filename = documento.get("fileName", "").lower()
        mime     = documento.get("mimeType", "").lower()
        if not url:
            return None
        if "pdf" in mime or filename.endswith(".pdf"):
            return await extrair_texto_pdf(url)
        if any(x in mime for x in ["excel", "spreadsheet", "xlsx", "xls"]) or \
           filename.endswith((".xlsx", ".xls")):
            return await extrair_texto_excel(url)
        return f"[Arquivo recebido: {filename} — tipo nao suportado para leitura automatica]"
    return None

# ============================================================
# CSS DO PAINEL
# ============================================================
CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: Arial, sans-serif; background: #f0f2f5; color: #333; }
header { background: #1a1a2e; color: white; padding: 16px 24px; display: flex; align-items: center; gap: 12px; }
header h1 { font-size: 18px; }
header a { color: #aab4ff; text-decoration: none; font-size: 14px; margin-left: auto; }
header a:hover { text-decoration: underline; }
.container { max-width: 860px; margin: 30px auto; padding: 0 20px; }
.card { background: white; border-radius: 12px; padding: 20px; margin-bottom: 16px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }
.card h2 { font-size: 16px; margin-bottom: 16px; color: #555; }
a { color: #4f46e5; text-decoration: none; }
a:hover { text-decoration: underline; }
.badge { display: inline-block; border-radius: 20px; padding: 2px 10px; font-size: 12px; margin-left: 8px; }
.badge-premium { background: #fef9c3; color: #b45309; }
.badge-freemium { background: #f1f5f9; color: #64748b; }
.badge-ativo { background: #dcfce7; color: #16a34a; }
.badge-inativo { background: #fee2e2; color: #dc2626; }
.badge-pendente { background: #fef3c7; color: #d97706; }
.aluno-row { display: flex; justify-content: space-between; align-items: center; padding: 12px 0; border-bottom: 1px solid #f0f0f0; }
.aluno-row:last-child { border-bottom: none; }
.aluno-info { font-size: 12px; color: #999; margin-top: 4px; max-width: 540px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.btn { display: inline-block; padding: 6px 14px; border-radius: 8px; font-size: 13px; cursor: pointer; border: none; text-decoration: none; }
.btn-primary { background: #4f46e5; color: white; }
.btn-primary:hover { background: #4338ca; text-decoration: none; color: white; }
.btn-success { background: #dcfce7; color: #16a34a; }
.btn-success:hover { background: #bbf7d0; text-decoration: none; }
.btn-danger { background: #fee2e2; color: #dc2626; }
.btn-danger:hover { background: #fecaca; text-decoration: none; }
.btn-warning { background: #fef3c7; color: #d97706; }
.btn-warning:hover { background: #fde68a; text-decoration: none; }
.back { display: inline-block; margin-bottom: 16px; font-size: 14px; }
.total { font-size: 13px; color: #888; margin-bottom: 16px; }
.chat { display: flex; flex-direction: column; gap: 10px; }
.msg { display: flex; flex-direction: column; max-width: 80%; }
.msg.usuario { align-self: flex-end; align-items: flex-end; }
.msg.agente { align-self: flex-start; align-items: flex-start; }
.label { font-size: 11px; color: #aaa; margin-bottom: 3px; padding: 0 6px; }
.balao { padding: 10px 14px; border-radius: 16px; font-size: 14px; line-height: 1.6; }
.usuario .balao { background: #dcf8c6; border-bottom-right-radius: 4px; }
.agente .balao { background: #f8f8f8; border-bottom-left-radius: 4px; box-shadow: 0 1px 2px rgba(0,0,0,0.08); }
textarea { width: 100%; padding: 12px; border: 1px solid #ddd; border-radius: 8px; font-size: 14px; font-family: monospace; line-height: 1.6; resize: vertical; min-height: 300px; }
textarea:focus { outline: none; border-color: #4f46e5; box-shadow: 0 0 0 2px rgba(79,70,229,0.1); }
input[type=text], input[type=number] { width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 8px; font-size: 14px; }
.success { background: #dcfce7; color: #16a34a; padding: 10px 16px; border-radius: 8px; margin-bottom: 16px; font-size: 14px; }
.nav { display: flex; gap: 12px; margin-bottom: 20px; flex-wrap: wrap; }
.nav a { padding: 8px 16px; border-radius: 8px; background: white; font-size: 14px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }
.nav a.ativo { background: #4f46e5; color: white; }
.nav a:hover { text-decoration: none; background: #e0e7ff; }
.nav a.ativo:hover { background: #4338ca; }
.stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 12px; margin-bottom: 20px; }
.stat { background: white; border-radius: 12px; padding: 16px; text-align: center; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }
.stat .num { font-size: 28px; font-weight: bold; color: #4f46e5; }
.stat .label { font-size: 12px; color: #888; margin-top: 4px; }
"""

def base_html(titulo: str, conteudo: str, pagina_ativa: str = "") -> str:
    nav = {
        "usuarios":    ("Usuarios",    "/admin"),
        "assinaturas": ("Assinaturas", "/admin/assinaturas"),
        "agenda":      ("Agenda",      "/admin/agenda"),
        "consultas":   ("Consultas",   "/admin/consultas"),
        "prompt":      ("Prompt",      "/admin/prompt"),
        "arquivos":    ("Arquivos",    "/admin/arquivos"),
    }
    nav_html = ""
    for chave, (label, url) in nav.items():
        ativo = 'class="ativo"' if pagina_ativa == chave else ""
        nav_html += f'<a href="{url}" {ativo}>{label}</a>'

    return f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{titulo} — {AGENT_NAME}</title>
    <style>{CSS}</style>
</head>
<body>
    <header>
        <span>🧠</span>
        <h1>{AGENT_NAME} — Painel Admin</h1>
        <a href="/admin">Inicio</a>
    </header>
    <div class="container">
        <div class="nav">{nav_html}</div>
        {conteudo}
    </div>
</body>
</html>"""

# ============================================================
# FUNCOES AUXILIARES
# ============================================================

async def enviar_whatsapp(telefone: str, mensagem: str):
    numero_limpo = telefone.replace("+", "").replace("-", "").replace(" ", "")
    if numero_limpo.startswith("55") and len(numero_limpo) == 12:
        numero_limpo = numero_limpo[:4] + "9" + numero_limpo[4:]
    url     = f"https://api.z-api.io/instances/{ZAPI_INSTANCE_ID}/token/{ZAPI_TOKEN}/send-text"
    headers = {"Content-Type": "application/json", "Client-Token": ZAPI_CLIENT_TOKEN}
    payload = {"phone": numero_limpo, "message": mensagem}
    print(f"ENVIANDO para {numero_limpo}")
    async with httpx.AsyncClient(timeout=30) as http:
        response = await http.post(url, headers=headers, json=payload)
        print(f"Z-API STATUS: {response.status_code} | {response.text}")


def obter_link_pagamento(telefone: str) -> str:
    """Gera link de pagamento do Mercado Pago com telefone como referencia."""
    if LINK_PAGAMENTO:
        return LINK_PAGAMENTO
    # Link direto para o plano MP com referencia do usuario
    return f"https://www.mercadopago.com.br/subscriptions/checkout?preapproval_plan_id={MP_PLAN_ID}&back_url={BASE_URL}/pagamento/obrigado&external_reference={telefone}"


async def chamar_claude(telefone: str, mensagem_usuario: str) -> str:
    # Detecta interesse em consulta na mensagem
    palavras_consulta = ["consulta", "agendar", "teleconsulta", "atendimento", "marcar"]
    if any(p in mensagem_usuario.lower() for p in palavras_consulta):
        # Verifica se ja tem nome registrado para consulta
        dados_consulta = r.get(f"{CONSULTA_PREFIX}{telefone}")
        if not dados_consulta:
            # Registra com nome desconhecido por enquanto, sera atualizado
            registrar_interesse_consulta(telefone, "Nome nao informado")
            print(f"INTERESSE CONSULTA registrado: {telefone}")

    salvar_mensagem(telefone, "user", mensagem_usuario)
    historico = obter_historico(telefone)

    hoje       = datetime.now().strftime("%d/%m/%Y")
    dia_semana = ["segunda-feira","terca-feira","quarta-feira","quinta-feira",
                  "sexta-feira","sabado","domingo"][datetime.now().weekday()]

    # Injeta status de assinatura no prompt
    status_usuario = "PREMIUM" if eh_premium(telefone) else "FREEMIUM"
    link_pg = obter_link_pagamento(telefone)

    prompt_base = injetar_arquivos_no_prompt(obter_prompt())
    prompt_base = prompt_base.replace("{STATUS}", status_usuario)
    prompt_base = prompt_base.replace("[LINK_PAGAMENTO]", link_pg)

    system = prompt_base + f"\n\nDATA ATUAL: {dia_semana}, {hoje}\nSTATUS DO USUARIO: {status_usuario}\nLINK DE PAGAMENTO: {link_pg}"

    resposta = client.messages.create(
        model=AGENT_MODEL,
        max_tokens=1024,
        system=system,
        messages=historico
    )

    texto_resposta = resposta.content[0].text
    salvar_mensagem(telefone, "assistant", texto_resposta)

    # Detecta se o bot mencionou interesse em consulta na resposta
    if "registrar seu interesse" in texto_resposta.lower() or "vou registrar" in texto_resposta.lower():
        # Tenta extrair nome da ultima mensagem do usuario
        registrar_interesse_consulta(telefone, mensagem_usuario[:50])

    return texto_resposta

# ============================================================
# ROTAS PUBLICAS
# ============================================================

@app.get("/")
def status():
    return {"status": f"{AGENT_NAME} online"}


@app.get("/pagamento", response_class=HTMLResponse)
async def pagina_pagamento(ref: str = ""):
    """Pagina de pagamento com link para o plano MP."""
    telefone = ref or "visitante"
    link = obter_link_pagamento(telefone)
    return HTMLResponse(f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Assinar PrimeiraMente Premium</title>
    <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ font-family: Arial, sans-serif; background: #f0f9ff; display: flex; align-items: center; justify-content: center; min-height: 100vh; padding: 20px; }}
        .card {{ background: white; border-radius: 16px; padding: 40px; max-width: 460px; width: 100%; box-shadow: 0 4px 24px rgba(0,0,0,0.10); text-align: center; }}
        .emoji {{ font-size: 48px; margin-bottom: 16px; }}
        h1 {{ font-size: 22px; color: #1a1a2e; margin-bottom: 8px; }}
        p {{ font-size: 15px; color: #555; margin-bottom: 24px; line-height: 1.6; }}
        .preco {{ font-size: 36px; font-weight: bold; color: #4f46e5; margin-bottom: 8px; }}
        .preco span {{ font-size: 16px; color: #888; font-weight: normal; }}
        ul {{ text-align: left; margin: 16px 0 28px; padding: 0 8px; list-style: none; }}
        ul li {{ padding: 6px 0; font-size: 14px; color: #444; }}
        ul li::before {{ content: "✓ "; color: #16a34a; font-weight: bold; }}
        .btn {{ display: block; background: #4f46e5; color: white; padding: 16px; border-radius: 10px; font-size: 16px; font-weight: bold; text-decoration: none; }}
        .btn:hover {{ background: #4338ca; }}
        .seguro {{ font-size: 12px; color: #aaa; margin-top: 16px; }}
    </style>
</head>
<body>
    <div class="card">
        <div class="emoji">🧠</div>
        <h1>PrimeiraMente Premium</h1>
        <p>Orientacao especializada em psiquiatria infantil para pais que querem entender e ajudar seus filhos de verdade.</p>
        <div class="preco">R$ 29,90 <span>/ mes</span></div>
        <ul>
            <li>Orientacao personalizada por comportamento</li>
            <li>Identificacao de sinais de TDAH e TEA</li>
            <li>Estrategias praticas baseadas em ciencia</li>
            <li>Como se preparar para consultas medicas</li>
            <li>Acesso ilimitado 24h pelo WhatsApp</li>
        </ul>
        <a href="{link}" class="btn">Assinar agora — R$ 29,90/mes</a>
        <p class="seguro">🔒 Pagamento seguro via Mercado Pago. Cancele quando quiser.</p>
    </div>
</body>
</html>""")


@app.get("/pagamento/obrigado", response_class=HTMLResponse)
async def pagamento_obrigado(external_reference: str = "", collection_status: str = ""):
    """Pagina de retorno apos pagamento no MP."""
    if collection_status == "approved" and external_reference:
        dados = {
            "telefone": external_reference,
            "status": "ativo",
            "plano": "premium",
            "origem": "mercadopago_retorno",
            "data_inicio": datetime.now().isoformat(),
            "expira": (datetime.now() + timedelta(days=35)).isoformat()
        }
        salvar_assinatura(external_reference, dados)
        await enviar_whatsapp(external_reference,
            "Seu acesso Premium esta ativo! 🎉\n\nBem-vindo ao PrimeiraMente Premium. Pode me contar agora o que esta acontecendo com seu filho — estou aqui para ajudar com orientacoes completas 🧠💙")

    return HTMLResponse("""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <title>Pagamento confirmado</title>
    <style>
        body { font-family: Arial, sans-serif; background: #f0fdf4; display: flex; align-items: center; justify-content: center; min-height: 100vh; }
        .card { background: white; border-radius: 16px; padding: 40px; max-width: 400px; text-align: center; box-shadow: 0 4px 24px rgba(0,0,0,0.08); }
        .emoji { font-size: 56px; margin-bottom: 16px; }
        h1 { color: #16a34a; margin-bottom: 12px; }
        p { color: #555; line-height: 1.6; }
    </style>
</head>
<body>
    <div class="card">
        <div class="emoji">🎉</div>
        <h1>Pagamento confirmado!</h1>
        <p>Seu acesso Premium ja esta ativo.<br>Volte ao WhatsApp — seu assistente esta te esperando!</p>
    </div>
</body>
</html>""")


# ============================================================
# PAGINA PUBLICA — AGENDAMENTO DE CONSULTA
# ============================================================

@app.get("/consulta", response_class=HTMLResponse)
async def pagina_consulta(tel: str = ""):
    """Pagina publica de agendamento de teleconsulta."""
    config   = obter_config_agenda()
    valor    = config.get("valor", VALOR_CONSULTA)
    datas    = gerar_proximas_datas(30)

    # Monta o calendario HTML
    cards_datas = ""
    for d in datas[:14]:  # exibe ate 14 dias
        slots_html = ""
        for slot in d["slots"]:
            ag_id_enc = f"{d['data']}|{slot}|{tel}"
            slots_html += f"""
            <button class="slot" onclick="selecionarSlot('{d['data']}', '{slot}', '{d['data_br']}')">{slot}</button>"""
        cards_datas += f"""
        <div class="dia-card">
            <div class="dia-header"><span class="dia-semana">{d['dia_semana']}</span><span class="dia-num">{d['data_br']}</span></div>
            <div class="slots">{slots_html}</div>
        </div>"""

    if not cards_datas:
        cards_datas = "<p style='color:#888;padding:20px;text-align:center'>Nenhum horario disponivel no momento. Entre em contato pelo WhatsApp.</p>"

    return HTMLResponse(f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Agendar Teleconsulta — Dra. Larissa</title>
    <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ font-family: 'Georgia', serif; background: #fafaf8; color: #2d2d2d; }}
        .hero {{ background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%); color: white; padding: 48px 24px; text-align: center; }}
        .hero .tag {{ display: inline-block; background: rgba(255,255,255,0.15); border-radius: 20px; padding: 4px 16px; font-size: 13px; margin-bottom: 16px; letter-spacing: 1px; }}
        .hero h1 {{ font-size: 28px; margin-bottom: 12px; font-weight: normal; }}
        .hero p {{ font-size: 15px; opacity: 0.85; max-width: 500px; margin: 0 auto; line-height: 1.7; }}
        .info-bar {{ display: flex; justify-content: center; gap: 32px; background: white; padding: 20px; border-bottom: 1px solid #eee; flex-wrap: wrap; }}
        .info-item {{ text-align: center; }}
        .info-item .val {{ font-size: 20px; font-weight: bold; color: #1a1a2e; }}
        .info-item .lab {{ font-size: 12px; color: #888; margin-top: 2px; }}
        .container {{ max-width: 780px; margin: 32px auto; padding: 0 20px; }}
        h2 {{ font-size: 18px; font-weight: normal; color: #444; margin-bottom: 20px; }}
        .calendario {{ display: flex; gap: 12px; overflow-x: auto; padding-bottom: 12px; }}
        .dia-card {{ min-width: 120px; background: white; border-radius: 12px; padding: 14px 10px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }}
        .dia-header {{ text-align: center; margin-bottom: 10px; border-bottom: 1px solid #f0f0f0; padding-bottom: 8px; }}
        .dia-semana {{ display: block; font-size: 11px; color: #888; text-transform: uppercase; letter-spacing: 1px; }}
        .dia-num {{ display: block; font-size: 13px; color: #333; margin-top: 2px; }}
        .slots {{ display: flex; flex-direction: column; gap: 6px; }}
        .slot {{ background: #f0f4ff; color: #4f46e5; border: none; border-radius: 8px; padding: 8px; font-size: 13px; cursor: pointer; transition: all 0.15s; }}
        .slot:hover {{ background: #4f46e5; color: white; }}
        .slot.selecionado {{ background: #4f46e5; color: white; }}
        .confirmacao {{ display: none; background: white; border-radius: 16px; padding: 28px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); margin-top: 24px; }}
        .confirmacao h3 {{ margin-bottom: 16px; color: #1a1a2e; }}
        .resumo {{ background: #f8f9ff; border-radius: 10px; padding: 16px; margin-bottom: 20px; }}
        .resumo-linha {{ display: flex; justify-content: space-between; padding: 6px 0; font-size: 14px; border-bottom: 1px solid #eee; }}
        .resumo-linha:last-child {{ border: none; font-weight: bold; color: #4f46e5; font-size: 16px; }}
        input[type=text], input[type=tel] {{ width: 100%; padding: 12px; border: 1px solid #ddd; border-radius: 8px; font-size: 14px; margin-bottom: 12px; font-family: Arial, sans-serif; }}
        input:focus {{ outline: none; border-color: #4f46e5; box-shadow: 0 0 0 2px rgba(79,70,229,0.1); }}
        .btn-pagar {{ display: block; width: 100%; background: #4f46e5; color: white; padding: 16px; border-radius: 10px; font-size: 16px; font-weight: bold; border: none; cursor: pointer; margin-top: 8px; }}
        .btn-pagar:hover {{ background: #4338ca; }}
        .seguro {{ font-size: 12px; color: #aaa; text-align: center; margin-top: 12px; }}
        .sobre {{ background: white; border-radius: 16px; padding: 28px; margin-top: 24px; box-shadow: 0 1px 4px rgba(0,0,0,0.06); }}
        .sobre h3 {{ color: #1a1a2e; margin-bottom: 12px; font-size: 16px; }}
        .sobre ul {{ list-style: none; padding: 0; }}
        .sobre ul li {{ padding: 6px 0; font-size: 14px; color: #555; }}
        .sobre ul li::before {{ content: "✓ "; color: #16a34a; font-weight: bold; }}
    </style>
</head>
<body>
    <div class="hero">
        <div class="tag">🧠 TELECONSULTA ONLINE</div>
        <h1>Dra. Larissa Alves Lourenco</h1>
        <p>Psiquiatra Infantil especializada em TDAH, TEA e desenvolvimento. CRM SP 223841</p>
    </div>
    <div class="info-bar">
        <div class="info-item"><div class="val">45 min</div><div class="lab">Duracao</div></div>
        <div class="info-item"><div class="val">Online</div><div class="lab">Modalidade</div></div>
        <div class="info-item"><div class="val">Individual</div><div class="lab">Formato</div></div>
        <div class="info-item"><div class="val">R$ {valor:.2f}</div><div class="lab">Valor</div></div>
    </div>
    <div class="container">
        <h2>Escolha um horario disponivel:</h2>
        <div class="calendario">
            {cards_datas}
        </div>

        <div class="confirmacao" id="confirmacao">
            <h3>Confirmar agendamento</h3>
            <div class="resumo">
                <div class="resumo-linha"><span>Consulta</span><span>Teleconsulta individual</span></div>
                <div class="resumo-linha"><span>Data</span><span id="res-data">—</span></div>
                <div class="resumo-linha"><span>Horario</span><span id="res-hora">—</span></div>
                <div class="resumo-linha"><span>Duracao</span><span>{DURACAO_CONSULTA} minutos</span></div>
                <div class="resumo-linha"><span>Total</span><span>R$ {valor:.2f}</span></div>
            </div>
            <input type="text" id="nome" placeholder="Seu nome completo" required>
            <input type="tel" id="telefone" placeholder="WhatsApp (com DDD, ex: 11999998888)" value="{tel}">
            <button class="btn-pagar" onclick="irParaPagamento()">Pagar e confirmar agendamento</button>
            <p class="seguro">🔒 Pagamento seguro via Mercado Pago</p>
        </div>

        <div class="sobre">
            <h3>O que voce pode esperar da consulta:</h3>
            <ul>
                <li>Avaliacao comportamental e orientacao especializada</li>
                <li>Discussao sobre sinais de TDAH, TEA ou outros transtornos</li>
                <li>Estrategias praticas para o dia a dia em familia</li>
                <li>Orientacao sobre quando e como buscar avaliacao presencial</li>
                <li>Link de videoconferencia enviado por WhatsApp apos pagamento</li>
            </ul>
        </div>
    </div>
    <script>
        let dataSelecionada = "", horaSelecionada = "", dataBR = "";

        function selecionarSlot(data, hora, dataBrStr) {{
            document.querySelectorAll('.slot').forEach(b => b.classList.remove('selecionado'));
            event.target.classList.add('selecionado');
            dataSelecionada = data;
            horaSelecionada = hora;
            dataBR = dataBrStr;
            document.getElementById('res-data').textContent = dataBrStr;
            document.getElementById('res-hora').textContent = hora;
            document.getElementById('confirmacao').style.display = 'block';
            document.getElementById('confirmacao').scrollIntoView({{behavior:'smooth'}});
        }}

        function irParaPagamento() {{
            const nome = document.getElementById('nome').value.trim();
            const tel  = document.getElementById('telefone').value.trim();
            if (!nome) {{ alert('Por favor, informe seu nome.'); return; }}
            if (!tel)  {{ alert('Por favor, informe seu WhatsApp.'); return; }}
            if (!dataSelecionada) {{ alert('Selecione um horario.'); return; }}
            const ref = encodeURIComponent(dataSelecionada + '|' + horaSelecionada + '|' + tel + '|' + nome);
            window.location.href = '/consulta/checkout?ref=' + ref;
        }}
    </script>
</body>
</html>""")


@app.get("/consulta/checkout")
async def consulta_checkout(ref: str = ""):
    """Cria preferencia de pagamento MP e redireciona."""
    if not ref or not MP_ACCESS_TOKEN:
        return RedirectResponse(url="/consulta")

    try:
        partes = ref.split("|")
        data   = partes[0] if len(partes) > 0 else ""
        hora   = partes[1] if len(partes) > 1 else ""
        tel    = partes[2] if len(partes) > 2 else ""
        nome   = partes[3] if len(partes) > 3 else "Paciente"
    except Exception:
        return RedirectResponse(url="/consulta")

    # Verifica disponibilidade antes de cobrar
    if not horario_disponivel(data, hora):
        return HTMLResponse("<h2 style='font-family:Arial;padding:40px;text-align:center'>Ops! Este horario foi ocupado. <a href='/consulta'>Escolha outro</a></h2>")

    # Reserva o slot temporariamente
    ag_id = criar_agendamento_id(data, hora)
    ag_temp = {
        "id": ag_id,
        "data": data,
        "hora": hora,
        "telefone": tel,
        "nome": nome,
        "status": "pendente",
        "criado_em": datetime.now().isoformat()
    }
    salvar_agendamento(ag_temp)

    config = obter_config_agenda()
    valor  = config.get("valor", VALOR_CONSULTA)

    try:
        async with httpx.AsyncClient(timeout=30) as http:
            res = await http.post(
                "https://api.mercadopago.com/checkout/preferences",
                headers={
                    "Authorization": f"Bearer {MP_ACCESS_TOKEN}",
                    "Content-Type": "application/json"
                },
                json={
                    "items": [{
                        "title": f"Teleconsulta Dra. Larissa — {data} as {hora}",
                        "quantity": 1,
                        "unit_price": float(valor),
                        "currency_id": "BRL"
                    }],
                    "payer": {"name": nome},
                    "external_reference": ref,
                    "back_urls": {
                        "success": f"{BASE_URL}/consulta/obrigado",
                        "failure": f"{BASE_URL}/consulta",
                        "pending": f"{BASE_URL}/consulta/obrigado"
                    },
                    "auto_return": "approved",
                    "notification_url": f"{BASE_URL}/webhook/mercadopago"
                }
            )
            dados_mp = res.json()
            link_mp  = dados_mp.get("init_point", "")
            if link_mp:
                return RedirectResponse(url=link_mp)
    except Exception as e:
        print(f"ERRO ao criar preferencia MP: {e}")

    return HTMLResponse("<h2 style='font-family:Arial;padding:40px;text-align:center'>Erro ao iniciar pagamento. <a href='/consulta'>Tente novamente</a></h2>")


@app.get("/consulta/obrigado", response_class=HTMLResponse)
async def consulta_obrigado(external_reference: str = "", collection_status: str = "", payment_id: str = ""):
    """Retorno apos pagamento da consulta."""
    if collection_status == "approved" and external_reference:
        try:
            partes = external_reference.split("|")
            data   = partes[0]
            hora   = partes[1]
            tel    = partes[2]
            nome   = partes[3] if len(partes) > 3 else "Paciente"

            ag_id = criar_agendamento_id(data, hora)
            ag = obter_agendamento(ag_id) or {}
            ag.update({
                "id": ag_id, "data": data, "hora": hora,
                "telefone": tel, "nome": nome,
                "status": "confirmado",
                "payment_id": payment_id,
                "confirmado_em": datetime.now().isoformat()
            })
            salvar_agendamento(ag)

            data_br = datetime.strptime(data, "%Y-%m-%d").strftime("%d/%m/%Y")
            await enviar_whatsapp(tel,
                f"Consulta confirmada! 🎉\n\n"
                f"📅 Data: {data_br}\n"
                f"🕐 Horario: {hora}\n"
                f"⏱ Duracao: {DURACAO_CONSULTA} minutos\n"
                f"💻 Formato: Teleconsulta online\n\n"
                f"O link de videoconferencia sera enviado aqui no WhatsApp 30 minutos antes da consulta.\n\n"
                f"Qualquer duvida, e so me chamar aqui! 💙")
        except Exception as e:
            print(f"ERRO ao confirmar consulta: {e}")

    return HTMLResponse("""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <title>Consulta confirmada!</title>
    <style>
        body { font-family: Arial, sans-serif; background: #f0fdf4; display: flex; align-items: center; justify-content: center; min-height: 100vh; padding: 20px; }
        .card { background: white; border-radius: 16px; padding: 40px; max-width: 420px; text-align: center; box-shadow: 0 4px 24px rgba(0,0,0,0.08); }
        .emoji { font-size: 56px; margin-bottom: 16px; }
        h1 { color: #16a34a; margin-bottom: 12px; font-size: 22px; }
        p { color: #555; line-height: 1.7; font-size: 15px; }
    </style>
</head>
<body>
    <div class="card">
        <div class="emoji">🎉</div>
        <h1>Consulta confirmada!</h1>
        <p>O pagamento foi aprovado e sua consulta esta agendada.<br><br>
        Voce recebera uma mensagem no WhatsApp com todos os detalhes e o link da videoconferencia.</p>
    </div>
</body>
</html>""")


# ============================================================
# PAINEL ADMIN — AGENDA
# ============================================================

@app.get("/admin/agenda", response_class=HTMLResponse)
def painel_agenda(admin: str = Depends(verificar_admin), msg: str = ""):
    agendamentos = listar_agendamentos(apenas_futuros=False)
    config       = obter_config_agenda()
    aviso = f'<div class="success">{msg}</div>' if msg else ""

    futuros    = [a for a in agendamentos if a.get("status") == "confirmado" and a.get("data","") >= datetime.now().strftime("%Y-%m-%d")]
    pendentes  = [a for a in agendamentos if a.get("status") == "pendente"]
    passados   = [a for a in agendamentos if a.get("status") == "confirmado" and a.get("data","") < datetime.now().strftime("%Y-%m-%d")]
    cancelados = [a for a in agendamentos if a.get("status") == "cancelado"]

    def render_ag(lista, mostrar_acoes=True):
        if not lista:
            return "<p style='color:#888;padding:12px 0'>Nenhum.</p>"
        rows = ""
        for ag in lista:
            tel  = ag.get("telefone","—")
            nome = ag.get("nome","—")
            data = ag.get("data","")
            hora = ag.get("hora","—")
            ag_id = criar_agendamento_id(data, hora)
            try:
                data_br = datetime.strptime(data, "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                data_br = data
            status = ag.get("status","—")
            badge_cor = {"confirmado":"badge-ativo","pendente":"badge-pendente","cancelado":"badge-inativo"}.get(status,"")
            acoes = ""
            if mostrar_acoes:
                acoes = f'<a href="/admin/agenda/cancelar/{ag_id}" onclick="return confirm(\'Cancelar este agendamento?\')" class="btn btn-warning" style="margin-right:6px">Cancelar</a>'
                acoes += f'<a href="/admin/agenda/apagar/{ag_id}" onclick="return confirm(\'Apagar permanentemente?\')" class="btn btn-danger">Apagar</a>'
            rows += f"""
            <div class="aluno-row">
                <div>
                    <div><strong>{nome}</strong> <span class="badge {badge_cor}">{status}</span></div>
                    <div class="aluno-info">📅 {data_br} as {hora} | 📱 {tel}</div>
                </div>
                <div style="display:flex;gap:6px;flex-wrap:wrap">{acoes}</div>
            </div>"""
        return rows

    # Bloqueios ativos
    bloqueios_chaves = r.keys(f"{BLOQUEIO_PREFIX}*")
    bloqueios_html = ""
    for chave in sorted(bloqueios_chaves):
        bl = chave.replace(BLOQUEIO_PREFIX, "")
        bloqueios_html += f"""
        <div class="aluno-row">
            <div><div>{bl.replace('_dia',' — DIA INTEIRO').replace('_',' as ')}</div></div>
            <a href="/admin/agenda/desbloquear/{bl}" class="btn btn-success">Desbloquear</a>
        </div>"""
    if not bloqueios_html:
        bloqueios_html = "<p style='color:#888;padding:12px 0'>Nenhum bloqueio ativo.</p>"

    horarios_check = ""
    for h in ["08:00","09:00","10:00","11:00","12:00","13:00","14:00","15:00","16:00","17:00","18:00","19:00","20:00"]:
        checked = "checked" if h in config.get("horarios", HORARIOS_PADRAO) else ""
        horarios_check += f'<label style="margin-right:12px;font-size:14px"><input type="checkbox" name="horarios" value="{h}" {checked}> {h}</label>'

    conteudo = f"""
    {aviso}
    <div class="stats">
        <div class="stat"><div class="num">{len(futuros)}</div><div class="label">Proximas</div></div>
        <div class="stat"><div class="num">{len(pendentes)}</div><div class="label">Pendentes</div></div>
        <div class="stat"><div class="num">{len(passados)}</div><div class="label">Realizadas</div></div>
    </div>

    <div class="card">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;">
            <h2>Proximas consultas ({len(futuros)})</h2>
            <a href="/admin/agenda/nova" class="btn btn-primary">+ Novo agendamento</a>
        </div>
        {render_ag(futuros)}
    </div>

    <div class="card">
        <h2>Pendentes de pagamento ({len(pendentes)})</h2>
        {render_ag(pendentes)}
    </div>

    <div class="card">
        <h2>Bloquear data ou horario</h2>
        <form method="post" action="/admin/agenda/bloquear" style="display:flex;gap:10px;flex-wrap:wrap;align-items:flex-end">
            <div>
                <label style="font-size:13px;color:#555;display:block;margin-bottom:4px">Data</label>
                <input type="date" name="data" required style="padding:10px;border:1px solid #ddd;border-radius:8px;font-size:14px">
            </div>
            <div>
                <label style="font-size:13px;color:#555;display:block;margin-bottom:4px">Horario (vazio = dia inteiro)</label>
                <input type="text" name="hora" placeholder="ex: 14:00" style="width:130px;padding:10px;border:1px solid #ddd;border-radius:8px;font-size:14px">
            </div>
            <button type="submit" class="btn btn-warning">Bloquear</button>
        </form>
        <div style="margin-top:16px">{bloqueios_html}</div>
    </div>

    <div class="card">
        <h2>Configurar horarios disponiveis</h2>
        <form method="post" action="/admin/agenda/config">
            <div style="margin-bottom:16px;line-height:2.2">{horarios_check}</div>
            <div style="margin-bottom:12px">
                <label style="font-size:13px;color:#555;display:block;margin-bottom:4px">Valor da consulta (R$)</label>
                <input type="number" name="valor" value="{config.get('valor', VALOR_CONSULTA)}" step="0.01" style="width:160px;padding:10px;border:1px solid #ddd;border-radius:8px;font-size:14px">
            </div>
            <button type="submit" class="btn btn-primary">Salvar configuracoes</button>
        </form>
    </div>

    <div class="card">
        <h2>Consultas realizadas ({len(passados)})</h2>
        {render_ag(passados, False)}
    </div>
    <div class="card">
        <h2>Canceladas ({len(cancelados)})</h2>
        {render_ag(cancelados, False)}
    </div>"""

    return HTMLResponse(base_html("Agenda", conteudo, "agenda"))


@app.get("/admin/agenda/nova", response_class=HTMLResponse)
def novo_agendamento_get(admin: str = Depends(verificar_admin)):
    conteudo = """
    <a class="back" href="/admin/agenda">← Voltar</a>
    <div class="card">
        <h2>Novo agendamento manual</h2>
        <form method="post" action="/admin/agenda/nova">
            <div style="margin-bottom:12px">
                <label style="font-size:13px;color:#555;display:block;margin-bottom:4px">Nome do paciente</label>
                <input type="text" name="nome" required placeholder="Maria Silva">
            </div>
            <div style="margin-bottom:12px">
                <label style="font-size:13px;color:#555;display:block;margin-bottom:4px">WhatsApp (com DDD)</label>
                <input type="text" name="telefone" required placeholder="5511999998888">
            </div>
            <div style="margin-bottom:12px;display:flex;gap:12px">
                <div style="flex:1">
                    <label style="font-size:13px;color:#555;display:block;margin-bottom:4px">Data</label>
                    <input type="date" name="data" required style="width:100%;padding:10px;border:1px solid #ddd;border-radius:8px;font-size:14px">
                </div>
                <div style="flex:1">
                    <label style="font-size:13px;color:#555;display:block;margin-bottom:4px">Horario</label>
                    <input type="text" name="hora" required placeholder="14:00" style="width:100%;padding:10px;border:1px solid #ddd;border-radius:8px;font-size:14px">
                </div>
            </div>
            <button type="submit" class="btn btn-primary">Criar agendamento</button>
        </form>
    </div>"""
    return HTMLResponse(base_html("Novo Agendamento", conteudo, "agenda"))


@app.post("/admin/agenda/nova")
async def novo_agendamento_post(
    nome: str = Form(...),
    telefone: str = Form(...),
    data: str = Form(...),
    hora: str = Form(...),
    admin: str = Depends(verificar_admin)
):
    ag_id = criar_agendamento_id(data, hora)
    ag = {
        "id": ag_id, "data": data, "hora": hora,
        "telefone": telefone, "nome": nome,
        "status": "confirmado",
        "origem": "manual",
        "criado_em": datetime.now().isoformat()
    }
    salvar_agendamento(ag)
    return RedirectResponse(url="/admin/agenda?msg=Agendamento+criado!", status_code=303)


@app.get("/admin/agenda/cancelar/{ag_id:path}")
async def cancelar_ag(ag_id: str, admin: str = Depends(verificar_admin)):
    ag = obter_agendamento(ag_id)
    if ag:
        cancelar_agendamento(ag_id)
        tel = ag.get("telefone","")
        if tel:
            data_br = datetime.strptime(ag.get("data",""), "%Y-%m-%d").strftime("%d/%m/%Y") if ag.get("data") else ""
            await enviar_whatsapp(tel,
                f"Sua consulta do dia {data_br} as {ag.get('hora','')} foi cancelada.\n\nSe quiser reagendar, acesse: {BASE_URL}/consulta ou nos chame aqui pelo WhatsApp.")
    return RedirectResponse(url="/admin/agenda?msg=Agendamento+cancelado!")


@app.get("/admin/agenda/apagar/{ag_id:path}")
def apagar_ag(ag_id: str, admin: str = Depends(verificar_admin)):
    apagar_agendamento(ag_id)
    return RedirectResponse(url="/admin/agenda?msg=Agendamento+apagado!")


@app.post("/admin/agenda/bloquear")
async def bloquear_rota(
    data: str = Form(...),
    hora: str = Form(""),
    admin: str = Depends(verificar_admin)
):
    if hora.strip():
        bloquear_horario(data, hora.strip())
        msg = f"Horario+{hora}+em+{data}+bloqueado!"
    else:
        bloquear_dia(data)
        msg = f"Dia+{data}+bloqueado!"
    return RedirectResponse(url=f"/admin/agenda?msg={msg}", status_code=303)


@app.get("/admin/agenda/desbloquear/{chave:path}")
def desbloquear_rota(chave: str, admin: str = Depends(verificar_admin)):
    r.delete(f"{BLOQUEIO_PREFIX}{chave}")
    return RedirectResponse(url="/admin/agenda?msg=Desbloqueado!")


@app.post("/admin/agenda/config")
async def salvar_config_agenda_rota(
    request: Request,
    valor: float = Form(VALOR_CONSULTA),
    admin: str = Depends(verificar_admin)
):
    form = await request.form()
    horarios = form.getlist("horarios")
    config = obter_config_agenda()
    config["horarios"] = sorted(horarios)
    config["valor"]    = valor
    salvar_config_agenda(config)
    return RedirectResponse(url="/admin/agenda?msg=Configuracoes+salvas!", status_code=303)

# ============================================================
# WEBHOOK MERCADO PAGO
# ============================================================

@app.post("/webhook/mercadopago")
async def webhook_mercadopago(request: Request):
    """Recebe notificacoes de pagamento do Mercado Pago."""
    try:
        dados = await request.json()
        print(f"MP WEBHOOK: {json.dumps(dados)[:300]}")

        tipo  = dados.get("type") or dados.get("action", "")
        id_mp = dados.get("data", {}).get("id") or dados.get("id")

        if not id_mp:
            return {"status": "ignorado"}

        # Assinatura criada ou atualizada
        if "subscription" in tipo or "preapproval" in tipo:
            await processar_assinatura_mp(str(id_mp))
        # Pagamento avulso (cobranca recorrente)
        elif "payment" in tipo:
            await processar_pagamento_mp(str(id_mp))

        return {"status": "ok"}
    except Exception as e:
        print(f"ERRO webhook MP: {e}")
        return {"status": "erro", "detalhe": str(e)}


async def processar_assinatura_mp(preapproval_id: str):
    """Busca detalhes da assinatura no MP e ativa/desativa o usuario."""
    if not MP_ACCESS_TOKEN:
        return
    try:
        async with httpx.AsyncClient(timeout=30) as http:
            res = await http.get(
                f"https://api.mercadopago.com/preapproval/{preapproval_id}",
                headers={"Authorization": f"Bearer {MP_ACCESS_TOKEN}"}
            )
            if res.status_code != 200:
                print(f"MP ASSINATURA ERRO: {res.status_code}")
                return
            dados = res.json()

        telefone        = dados.get("external_reference", "")
        status_mp       = dados.get("status", "")  # authorized, paused, cancelled
        proximo_debito  = dados.get("next_payment_date", "")

        if not telefone:
            print("MP: external_reference vazio — nao foi possivel identificar usuario")
            return

        status_local = "ativo" if status_mp == "authorized" else "inativo"
        expira = None
        if proximo_debito:
            try:
                expira = (datetime.fromisoformat(proximo_debito[:10]) + timedelta(days=5)).isoformat()
            except Exception:
                expira = (datetime.now() + timedelta(days=35)).isoformat()

        assinatura = {
            "telefone": telefone,
            "status": status_local,
            "plano": "premium",
            "preapproval_id": preapproval_id,
            "status_mp": status_mp,
            "data_inicio": datetime.now().isoformat(),
            "expira": expira or (datetime.now() + timedelta(days=35)).isoformat()
        }
        salvar_assinatura(telefone, assinatura)
        print(f"ASSINATURA ATUALIZADA: {telefone} → {status_local}")

        if status_local == "ativo":
            await enviar_whatsapp(telefone,
                "Seu acesso Premium esta ativo! 🎉\n\nBem-vindo ao PrimeiraMente Premium. Pode me contar o que esta acontecendo com seu filho — estou aqui para ajudar com orientacoes completas 🧠💙")
        elif status_mp in ("cancelled", "paused"):
            await enviar_whatsapp(telefone,
                "Seu plano Premium foi cancelado. Sentiremos sua falta 💙\n\nSe quiser reativar a qualquer momento, e so me chamar aqui!")

    except Exception as e:
        print(f"ERRO ao processar assinatura MP: {e}")


async def processar_pagamento_mp(payment_id: str):
    """Processa um pagamento recorrente aprovado."""
    if not MP_ACCESS_TOKEN:
        return
    try:
        async with httpx.AsyncClient(timeout=30) as http:
            res = await http.get(
                f"https://api.mercadopago.com/v1/payments/{payment_id}",
                headers={"Authorization": f"Bearer {MP_ACCESS_TOKEN}"}
            )
            if res.status_code != 200:
                return
            dados = res.json()

        status_pg  = dados.get("status", "")
        telefone   = dados.get("external_reference", "")

        if not telefone:
            return

        if status_pg == "approved":
            assinatura = obter_assinatura(telefone)
            assinatura["status"] = "ativo"
            assinatura["expira"] = (datetime.now() + timedelta(days=35)).isoformat()
            assinatura["ultimo_pagamento"] = datetime.now().isoformat()
            salvar_assinatura(telefone, assinatura)
            print(f"PAGAMENTO APROVADO: {telefone}")
        elif status_pg in ("rejected", "cancelled"):
            assinatura = obter_assinatura(telefone)
            assinatura["status"] = "inativo"
            salvar_assinatura(telefone, assinatura)
            await enviar_whatsapp(telefone,
                "Tivemos um problema com o pagamento da sua assinatura 😕\n\nPor favor, atualize seu metodo de pagamento para continuar com o acesso Premium.")

    except Exception as e:
        print(f"ERRO ao processar pagamento MP: {e}")

# ============================================================
# PAINEL ADMIN — USUARIOS
# ============================================================

@app.get("/admin", response_class=HTMLResponse)
def painel_admin(admin: str = Depends(verificar_admin)):
    chaves = r.keys("historico:*")

    total_usuarios  = len(chaves)
    total_premium   = len([1 for c in r.keys(f"{ASSINATURA_PREFIX}*")
                           if json.loads(r.get(c) or "{}").get("status") == "ativo"])
    total_consultas = len([1 for c in r.keys(f"{CONSULTA_PREFIX}*")
                           if not json.loads(r.get(c) or "{}").get("atendido")])

    stats = f"""
    <div class="stats">
        <div class="stat"><div class="num">{total_usuarios}</div><div class="label">Usuarios</div></div>
        <div class="stat"><div class="num">{total_premium}</div><div class="label">Premium ativos</div></div>
        <div class="stat"><div class="num">{total_consultas}</div><div class="label">Consultas pendentes</div></div>
    </div>"""

    rows = ""
    for chave in sorted(chaves):
        telefone  = chave.replace("historico:", "")
        historico = obter_historico(telefone)
        total     = len(historico)
        ultima    = historico[-1]["content"][:80] + "..." if historico else "—"
        premium   = eh_premium(telefone)
        badge     = '<span class="badge badge-premium">Premium</span>' if premium else '<span class="badge badge-freemium">Freemium</span>'
        rows += f"""
        <div class="aluno-row">
            <div>
                <div>
                    <a href="/admin/conversa/{telefone}">{telefone}</a>
                    {badge}
                    <span class="badge" style="background:#e0e7ff;color:#4f46e5">{total} msgs</span>
                </div>
                <div class="aluno-info">{ultima}</div>
            </div>
            <a href="/admin/apagar/{telefone}" onclick="return confirm('Apagar historico de {telefone}?')">
                <span class="btn btn-danger">Apagar</span>
            </a>
        </div>"""

    if not rows:
        rows = "<p style='color:#888;padding:20px 0'>Nenhum usuario ainda.</p>"

    conteudo = stats + f"""
    <div class="card">
        <h2>Usuarios ({total_usuarios} total)</h2>
        {rows}
    </div>"""

    return HTMLResponse(base_html("Usuarios", conteudo, "usuarios"))


@app.get("/admin/conversa/{telefone}", response_class=HTMLResponse)
def ver_conversa(telefone: str, admin: str = Depends(verificar_admin)):
    historico = obter_historico(telefone)
    assinatura = obter_assinatura(telefone)
    premium = eh_premium(telefone)

    status_badge = '<span class="badge badge-ativo">Premium ativo</span>' if premium else '<span class="badge badge-freemium">Freemium</span>'
    expira = assinatura.get("expira", "")
    expira_txt = f" — expira {expira[:10]}" if expira and premium else ""

    if not historico:
        conteudo = f"""
        <a class="back" href="/admin">← Voltar</a>
        <div class="card"><p>Nenhuma conversa para {telefone}.</p></div>"""
        return HTMLResponse(base_html(telefone, conteudo))

    msgs = ""
    for msg in historico:
        role  = msg["role"]
        texto = msg["content"].replace("\n", "<br>")
        if role == "user":
            msgs += f'<div class="msg usuario"><div class="label">Usuario</div><div class="balao">{texto}</div></div>'
        else:
            msgs += f'<div class="msg agente"><div class="label">{AGENT_NAME}</div><div class="balao">{texto}</div></div>'

    conteudo = f"""
    <a class="back" href="/admin">← Voltar</a>
    <div class="card" style="margin-bottom:12px;">
        <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px;">
            <div>{status_badge}{expira_txt}</div>
            <div style="display:flex;gap:8px;">
                {"" if premium else f'<a href="/admin/assinaturas/ativar/{telefone}" class="btn btn-success">Ativar Premium</a>'}
                {f'<a href="/admin/assinaturas/desativar/{telefone}" onclick="return confirm(\'Desativar {telefone}?\')" class="btn btn-warning">Desativar</a>' if premium else ""}
            </div>
        </div>
    </div>
    <div class="card">
        <h2>Conversa com {telefone}</h2>
        <div class="total">{len(historico)} mensagens</div>
        <div class="chat">{msgs}</div>
    </div>"""

    return HTMLResponse(base_html(telefone, conteudo))


@app.get("/admin/apagar/{telefone}")
def apagar_historico(telefone: str, admin: str = Depends(verificar_admin)):
    r.delete(f"historico:{telefone}")
    return RedirectResponse(url="/admin")

# ============================================================
# PAINEL ADMIN — ASSINATURAS
# ============================================================

@app.get("/admin/assinaturas", response_class=HTMLResponse)
def painel_assinaturas(admin: str = Depends(verificar_admin), msg: str = ""):
    assinaturas = listar_assinaturas()
    aviso = f'<div class="success">{msg}</div>' if msg else ""

    rows = ""
    for assin in assinaturas:
        tel     = assin.get("telefone", "")
        status  = assin.get("status", "freemium")
        plano   = assin.get("plano", "freemium")
        expira  = assin.get("expira", "")[:10] if assin.get("expira") else "—"
        origem  = assin.get("origem", assin.get("status_mp", "manual"))

        if status == "ativo":
            badge = '<span class="badge badge-ativo">Ativo</span>'
        elif status == "inativo":
            badge = '<span class="badge badge-inativo">Inativo</span>'
        else:
            badge = '<span class="badge badge-freemium">Freemium</span>'

        btns = ""
        if status != "ativo":
            btns += f'<a href="/admin/assinaturas/ativar/{tel}" class="btn btn-success" style="margin-right:6px">Ativar</a>'
        if status == "ativo":
            btns += f'<a href="/admin/assinaturas/desativar/{tel}" onclick="return confirm(\'Desativar {tel}?\')" class="btn btn-warning" style="margin-right:6px">Desativar</a>'
        btns += f'<a href="/admin/assinaturas/apagar/{tel}" onclick="return confirm(\'Remover {tel}?\')" class="btn btn-danger">Remover</a>'

        rows += f"""
        <div class="aluno-row">
            <div>
                <div><strong>{tel}</strong> {badge} <span class="badge" style="background:#e0e7ff;color:#4f46e5">{plano}</span></div>
                <div class="aluno-info">Expira: {expira} | Origem: {origem}</div>
            </div>
            <div style="display:flex;gap:6px;flex-wrap:wrap">{btns}</div>
        </div>"""

    if not rows:
        rows = "<p style='color:#888;padding:20px 0'>Nenhuma assinatura ainda. Elas aparecem aqui apos o primeiro pagamento ou ativacao manual.</p>"

    conteudo = f"""
    {aviso}
    <div class="card">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;">
            <h2>Assinaturas ({len(assinaturas)})</h2>
            <a href="/admin/assinaturas/nova" class="btn btn-primary">+ Ativar manualmente</a>
        </div>
        {rows}
    </div>"""

    return HTMLResponse(base_html("Assinaturas", conteudo, "assinaturas"))


@app.get("/admin/assinaturas/nova", response_class=HTMLResponse)
def nova_assinatura_get(admin: str = Depends(verificar_admin)):
    conteudo = """
    <a class="back" href="/admin/assinaturas">← Voltar</a>
    <div class="card">
        <h2>Ativar assinatura manualmente</h2>
        <p style="font-size:13px;color:#888;margin-bottom:16px;">
            Use para ativar um usuario que pagou por fora, por cortesia ou para testes.
        </p>
        <form method="post" action="/admin/assinaturas/nova">
            <div style="margin-bottom:12px;">
                <label style="font-size:13px;color:#555;display:block;margin-bottom:6px;">Telefone (com DDI, ex: 5511999998888)</label>
                <input type="text" name="telefone" required placeholder="5511999998888">
            </div>
            <div style="margin-bottom:12px;">
                <label style="font-size:13px;color:#555;display:block;margin-bottom:6px;">Dias de acesso</label>
                <input type="number" name="dias" value="30" min="1" max="365" style="width:120px;padding:10px;border:1px solid #ddd;border-radius:8px;">
            </div>
            <button type="submit" class="btn btn-primary">Ativar acesso</button>
        </form>
    </div>"""
    return HTMLResponse(base_html("Nova Assinatura", conteudo, "assinaturas"))


@app.post("/admin/assinaturas/nova")
async def nova_assinatura_post(
    telefone: str = Form(...),
    dias: int = Form(30),
    admin: str = Depends(verificar_admin)
):
    dados = {
        "telefone": telefone,
        "status": "ativo",
        "plano": "premium",
        "origem": "manual",
        "data_inicio": datetime.now().isoformat(),
        "expira": (datetime.now() + timedelta(days=dias)).isoformat()
    }
    salvar_assinatura(telefone, dados)
    print(f"ASSINATURA MANUAL: {telefone} por {dias} dias")
    return RedirectResponse(url="/admin/assinaturas?msg=Assinatura+ativada+com+sucesso!", status_code=303)


@app.get("/admin/assinaturas/ativar/{telefone}")
def ativar_assinatura(telefone: str, admin: str = Depends(verificar_admin)):
    assinatura = obter_assinatura(telefone)
    assinatura["status"] = "ativo"
    assinatura["plano"]  = "premium"
    assinatura["expira"] = (datetime.now() + timedelta(days=35)).isoformat()
    assinatura["origem"] = assinatura.get("origem", "manual")
    salvar_assinatura(telefone, assinatura)
    return RedirectResponse(url="/admin/assinaturas?msg=Usuario+ativado!")


@app.get("/admin/assinaturas/desativar/{telefone}")
def desativar_assinatura(telefone: str, admin: str = Depends(verificar_admin)):
    assinatura = obter_assinatura(telefone)
    assinatura["status"] = "inativo"
    salvar_assinatura(telefone, assinatura)
    return RedirectResponse(url="/admin/assinaturas?msg=Usuario+desativado!")


@app.get("/admin/assinaturas/apagar/{telefone}")
def apagar_assinatura(telefone: str, admin: str = Depends(verificar_admin)):
    r.delete(f"{ASSINATURA_PREFIX}{telefone}")
    return RedirectResponse(url="/admin/assinaturas")

# ============================================================
# PAINEL ADMIN — CONSULTAS
# ============================================================

@app.get("/admin/consultas", response_class=HTMLResponse)
def painel_consultas(admin: str = Depends(verificar_admin)):
    consultas = listar_consultas()
    pendentes = [c for c in consultas if not c.get("atendido")]
    atendidas = [c for c in consultas if c.get("atendido")]

    def render_rows(lista, mostrar_btn=True):
        if not lista:
            return "<p style='color:#888;padding:12px 0'>Nenhuma.</p>"
        rows = ""
        for c in lista:
            tel  = c.get("telefone", "")
            nome = c.get("nome", "—")
            data = c.get("data", "—")
            btn  = f'<a href="/admin/consultas/atender/{tel}" class="btn btn-success">Marcar como atendido</a>' if mostrar_btn else '<span style="font-size:12px;color:#16a34a">✓ Atendido</span>'
            rows += f"""
            <div class="aluno-row">
                <div>
                    <div><strong>{tel}</strong></div>
                    <div class="aluno-info">Mensagem: {nome} | Data: {data}</div>
                </div>
                {btn}
            </div>"""
        return rows

    conteudo = f"""
    <div class="card">
        <h2>Interessados em Consulta — Pendentes ({len(pendentes)})</h2>
        {render_rows(pendentes, True)}
    </div>
    <div class="card">
        <h2>Ja Atendidos ({len(atendidas)})</h2>
        {render_rows(atendidas, False)}
    </div>"""

    return HTMLResponse(base_html("Consultas", conteudo, "consultas"))


@app.get("/admin/consultas/atender/{telefone}")
def atender_consulta(telefone: str, admin: str = Depends(verificar_admin)):
    marcar_consulta_atendida(telefone)
    return RedirectResponse(url="/admin/consultas")

# ============================================================
# PAINEL ADMIN — EDITAR PROMPT
# ============================================================

@app.get("/admin/prompt", response_class=HTMLResponse)
def editar_prompt_get(admin: str = Depends(verificar_admin), salvo: str = ""):
    prompt_atual = obter_prompt()
    aviso = '<div class="success">Prompt salvo com sucesso!</div>' if salvo == "1" else ""
    conteudo = f"""
    {aviso}
    <div class="card">
        <h2>Editar Prompt do Agente</h2>
        <p style="font-size:13px;color:#888;margin-bottom:16px;">
            Use <code>{{STATUS}}</code> para injetar FREEMIUM ou PREMIUM dinamicamente.<br>
            Use <code>[LINK_PAGAMENTO]</code> para injetar o link de assinatura.
        </p>
        <form method="post" action="/admin/prompt">
            <textarea name="prompt">{prompt_atual}</textarea>
            <br><br>
            <button type="submit" class="btn btn-primary">Salvar Prompt</button>
        </form>
    </div>"""
    return HTMLResponse(base_html("Editar Prompt", conteudo, "prompt"))


@app.post("/admin/prompt")
async def editar_prompt_post(
    prompt: str = Form(...),
    admin: str = Depends(verificar_admin)
):
    salvar_prompt(prompt.strip())
    return RedirectResponse(url="/admin/prompt?salvo=1", status_code=303)

# ============================================================
# PAINEL ADMIN — ARQUIVOS DE REFERENCIA
# ============================================================

@app.get("/admin/arquivos", response_class=HTMLResponse)
async def painel_arquivos(admin: str = Depends(verificar_admin), salvo: str = ""):
    arquivos = listar_arquivos()
    aviso = '<div class="success">Arquivo salvo com sucesso!</div>' if salvo == "1" else ""

    rows = ""
    for arq in arquivos:
        kb = round(arq["tamanho"] / 1024, 1)
        nome_arq = arq["nome"]
        rows += f"""
        <div class="aluno-row">
            <div>
                <div><strong>[{nome_arq}]</strong> &nbsp;<span style="font-size:12px;color:#888;">{kb} KB</span></div>
                <div class="aluno-info">Use [{nome_arq}] no prompt para referenciar este arquivo</div>
            </div>
            <a href="/admin/arquivos/apagar/{nome_arq}" onclick="return confirm('Apagar {nome_arq}?')">
                <span class="btn btn-danger">Apagar</span>
            </a>
        </div>"""

    if not rows:
        rows = "<p style='color:#888;padding:12px 0'>Nenhum arquivo ainda.</p>"

    conteudo = f"""
    {aviso}
    <div class="card">
        <h2>Arquivos de Referencia</h2>
        <p style="font-size:13px;color:#888;margin-bottom:16px;">
            Faca upload de PDFs ou arquivos de texto. Depois use <code>[nome]</code> no prompt.
        </p>
        {rows}
    </div>
    <div class="card">
        <h2>Novo Arquivo</h2>
        <form method="post" action="/admin/arquivos" enctype="multipart/form-data">
            <div style="margin-bottom:12px;">
                <label style="font-size:13px;color:#555;display:block;margin-bottom:6px;">Nome de referencia (sem espacos)</label>
                <input type="text" name="nome" required placeholder="metodologia" style="width:100%;padding:10px;border:1px solid #ddd;border-radius:8px;font-size:14px;">
            </div>
            <div style="margin-bottom:16px;">
                <label style="font-size:13px;color:#555;display:block;margin-bottom:6px;">Arquivo (PDF ou TXT — max 500KB)</label>
                <input type="file" name="arquivo" accept=".pdf,.txt,.md" required style="font-size:14px;">
            </div>
            <button type="submit" class="btn btn-primary">Salvar Arquivo</button>
        </form>
    </div>"""

    return HTMLResponse(base_html("Arquivos", conteudo, "arquivos"))


@app.post("/admin/arquivos")
async def upload_arquivo(
    nome: str = Form(...),
    arquivo: UploadFile = File(...),
    admin: str = Depends(verificar_admin)
):
    conteudo_bytes = await arquivo.read()
    filename = arquivo.filename.lower()

    if filename.endswith(".pdf"):
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(conteudo_bytes)) as pdf:
                paginas = [p.extract_text() for p in pdf.pages[:30] if p.extract_text()]
            texto_final = "\n\n".join(paginas)
        except Exception as e:
            texto_final = f"[Erro ao ler PDF: {e}]"
    else:
        try:
            texto_final = conteudo_bytes.decode("utf-8")
        except Exception:
            texto_final = conteudo_bytes.decode("latin-1", errors="ignore")

    nome_seguro = re.sub(r"[^a-zA-Z0-9_\-]", "", nome).lower() or "arquivo"
    salvar_arquivo(nome_seguro, texto_final)
    return RedirectResponse(url="/admin/arquivos?salvo=1", status_code=303)


@app.get("/admin/arquivos/apagar/{nome}")
def apagar_arquivo_rota(nome: str, admin: str = Depends(verificar_admin)):
    apagar_arquivo(nome)
    return RedirectResponse(url="/admin/arquivos")

# ============================================================
# WEBHOOK Z-API (WhatsApp)
# ============================================================

@app.post("/webhook")
async def webhook(request: Request):
    try:
        dados = await request.json()

        if dados.get("type") != "ReceivedCallback":
            return {"status": "ignorado"}
        if dados.get("fromMe"):
            return {"status": "ignorado"}
        if dados.get("isGroup"):
            return {"status": "ignorado"}

        telefone = dados.get("phone", "")
        if not telefone:
            return {"status": "ignorado"}

        texto_midia = await processar_midia(dados)
        if texto_midia:
            resposta = await chamar_claude(telefone, texto_midia)
            await enviar_whatsapp(telefone, resposta)
            return {"status": "ok"}

        texto = dados.get("text", {}).get("message", "")
        if not texto:
            return {"status": "ignorado"}

        print(f"MSG de {telefone}: {texto[:80]}")
        resposta = await chamar_claude(telefone, texto)
        await enviar_whatsapp(telefone, resposta)
        return {"status": "ok"}

    except Exception as e:
        print(f"ERRO webhook: {e}")
        return {"status": "erro", "detalhe": str(e)}
